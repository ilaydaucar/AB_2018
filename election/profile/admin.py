from django.contrib import admin
from election.profile.models import UserProfile,City,Town
from openpyxl import load_workbook

def get_cities_and_towns():
    print("Dosya import edeceğim")
    path = "/Users/ilayda/Documents/City.xlsx"
    wb = load_workbook(path, data_only=True)
    sheet = wb.active
    for row in sheet.iter_rows(row_offset=1):
        code = str(row[0].value).strip()
        city_name = str(row[1].value).strip()
        town_name = str(row[2].value).strip()
        print(code, city_name, town_name)
        try:
            city = City.objects.get(code=code)
            town = Town(name=town_name, city=city)
            town.save()
        except:
            city = City(name=city_name, code=code)
            city.save()
            town = Town(name=town_name, city=city)
            town.save()


class CityAdmin(admin.ModelAdmin):
    list_display = ("name", "code",)


    def import_cities_and_towns(modeladmin, request, queryset):
        get_cities_and_towns()

    actions = [import_cities_and_towns, ]

    import_cities_and_towns.short_description = "İl ve ilçeleri Yükle"


class TownAdmin(admin.ModelAdmin):
    list_display = ("name", "city",)

class UserProfileAdmin(admin.ModelAdmin):
    list_display = ('email', 'name',  'is_active', 'is_staff',
                    'is_superuser', 'created_at', "town", 'city_name' )
    list_filter = ("town",)

    def city_name(self,obj):
        return obj.town.city.name

    def get_queryset(self, request):
        town = request.user.town
        users = UserProfile.objects.filter(town=town)
        return users




admin.site.register(UserProfile, UserProfileAdmin)
admin.site.register(City,CityAdmin)
admin.site.register(Town,TownAdmin)

